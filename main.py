from openpyxl import load_workbook
import pprint

print('================== WELCOME TO YOUR EXCEL AUTOMATION APP =======================')
print('\n'+' '*10+'Press any key to continue...')
input('> ')
print(' '*10 +'Opening Excel Spreadsheet...')
spreadsheet = load_workbook('transactions.xlsx')
sheet = spreadsheet.active
print(' '*10 +'Reading rows from the Excel spreadsheet...')
supplier_data = {}

for row_num in range(2,sheet.max_row):
    transaction_type = sheet[f'B{row_num}'].value
    supplier_name = sheet[f'C{row_num}'].value
    amount = sheet[f'D{row_num}'].value

    supplier_data.setdefault(transaction_type, {})
    supplier_data[transaction_type].setdefault(supplier_name, {'transaction_count': 0,
                                                               'amount':0 })
    supplier_data[transaction_type][supplier_name]['transaction_count']+=1
    supplier_data[transaction_type][supplier_name]['amount']+=int(amount)

with open('output.py', 'w') as ouput_file:
    ouput_file.write('\n')
    ouput_file.write('all_transactions = '+ pprint.pformat(supplier_data))

print(' '*10+'The Task Is Completed')
print(' '*10+'Check Your file, output.py. Have a great.')